"""
Generate CI midpoints/widths Excel for storage-enhanced rolling-window results.
Data sourced from writing/section_5_rolling_window.tex (tab:rolling_e01/e05/e10).
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# ---------------------------------------------------------------------------
# Raw CI data  {country: {date: (lo, hi) or None}}
# None = degenerate (SR = 1.000 exactly, shown as --- in table)
# ---------------------------------------------------------------------------
DATES = ["D-271", "D-301", "D-332", "D-362"]
DATE_LABELS = {
    "D-271": "D-271\n28 Sep",
    "D-301": "D-301\n28 Oct",
    "D-332": "D-332\n28 Nov",
    "D-362": "D-362\n28 Dec",
}

COUNTRIES = [
    "AT","BE","BG","CH","CZ","DE","DK","EE","ES","FI",
    "FR","GB","GR","HR","HU","IE","IT","LT","LU","LV",
    "NL","NO","PL","PT","RO","SE","SI","SK",
]

FOCUS = {"GR", "ES", "CH", "PT"}

# key: epsilon label
RAW = {
    "0.01": {
        "AT": [(1.307,1.369),(1.291,1.365),(1.320,1.416),(1.294,1.370)],
        "BE": [(1.260,1.340),(1.319,1.403),(1.448,1.594),(1.503,1.667)],
        "BG": [(1.254,1.334),(1.351,1.437),(1.276,1.340),(1.244,1.312)],
        "CH": [(1.074,1.100),(1.071,1.099),(1.064,1.096),(1.079,1.133)],
        "CZ": [(1.122,1.192),(1.107,1.145),(1.179,1.259),(1.307,1.437)],
        "DE": [(1.253,1.375),(1.226,1.302),(1.307,1.397),(1.531,1.781)],
        "DK": [(1.257,1.327),(1.264,1.320),(1.202,1.276),(1.077,1.173)],
        "EE": [(1.058,1.076),(1.054,1.076),(1.043,1.071),(1.016,1.042)],
        "ES": [(1.159,1.195),(1.141,1.225),(1.195,1.231),(1.185,1.269)],
        "FI": [(1.020,1.030),(1.022,1.032),(1.015,1.025),(1.009,1.013)],
        "FR": [(1.195,1.263),(1.282,1.364),(1.296,1.362),(1.213,1.307)],
        "GB": [(1.296,1.390),(1.240,1.336),(1.329,1.465),(1.275,1.455)],
        "GR": [(1.164,1.208),(1.149,1.197),(1.220,1.290),(1.190,1.242)],
        "HR": [(1.185,1.257),(1.257,1.305),(1.191,1.285),(1.143,1.191)],
        "HU": [(1.163,1.185),(1.204,1.252),(1.169,1.215),(1.149,1.187)],
        "IE": [None, None, None, (0.999,1.003)],
        "IT": [(1.106,1.148),(1.158,1.230),(1.188,1.248),(1.096,1.136)],
        "LT": [(1.319,1.425),(1.357,1.421),(1.322,1.452),(1.231,1.413)],
        "LU": [(1.235,1.305),(1.189,1.309),(1.263,1.381),(1.408,1.588)],
        "LV": [(1.069,1.105),(1.070,1.092),(1.077,1.105),(1.034,1.066)],
        "NL": [(1.372,1.526),(1.340,1.444),(1.393,1.525),(1.246,1.458)],
        "NO": [(1.013,1.029),(1.013,1.031),(1.011,1.019),(1.004,1.010)],
        "PL": [(1.054,1.070),(1.057,1.077),(1.040,1.060),(1.040,1.050)],
        "PT": [(1.130,1.190),(1.150,1.182),(1.132,1.212),(1.169,1.195)],
        "RO": [(1.326,1.400),(1.295,1.371),(1.376,1.480),(1.305,1.403)],
        "SE": [(1.065,1.081),(1.056,1.076),(1.064,1.078),(1.025,1.039)],
        "SI": [(1.010,1.016),(1.013,1.023),(1.022,1.034),(1.010,1.020)],
        "SK": [(1.004,1.010),(1.007,1.011),(1.007,1.013),(1.007,1.015)],
    },
    "0.05": {
        "AT": [(1.286,1.388),(1.234,1.342),(1.168,1.322),(1.158,1.252)],
        "BE": [(1.290,1.348),(1.357,1.437),(1.437,1.557),(1.414,1.490)],
        "BG": [(1.334,1.410),(1.367,1.415),(1.304,1.372),(1.263,1.333)],
        "CH": [(1.061,1.081),(1.065,1.091),(1.076,1.092),(1.085,1.105)],
        "CZ": [(1.134,1.180),(1.122,1.156),(1.183,1.239),(1.337,1.405)],
        "DE": [(1.276,1.342),(1.246,1.344),(1.323,1.437),(1.293,1.391)],
        "DK": [(1.203,1.293),(1.268,1.354),(1.166,1.264),(1.079,1.131)],
        "EE": [(1.061,1.073),(1.041,1.073),(1.022,1.040),(1.010,1.016)],
        "ES": [(1.112,1.190),(1.187,1.245),(1.209,1.255),(1.163,1.241)],
        "FI": [(1.017,1.033),(1.008,1.018),(1.011,1.019),(1.002,1.010)],
        "FR": [(1.219,1.271),(1.248,1.354),(1.302,1.388),(1.161,1.223)],
        "GB": [(1.187,1.251),(1.294,1.368),(1.284,1.496),(1.130,1.232)],
        "GR": [(1.193,1.247),(1.179,1.215),(1.289,1.351),(1.213,1.265)],
        "HR": [(1.170,1.248),(1.203,1.297),(1.188,1.274),(1.183,1.229)],
        "HU": [(1.188,1.230),(1.097,1.177),(1.145,1.197),(1.173,1.237)],
        "IE": [None, None, None, (0.997,1.001)],
        "IT": [(1.126,1.146),(1.163,1.213),(1.232,1.282),(1.127,1.155)],
        "LT": [(1.318,1.438),(1.344,1.440),(1.247,1.407),(1.129,1.243)],
        "LU": [(1.248,1.306),(1.236,1.292),(1.316,1.406),(1.448,1.546)],
        "LV": [(1.066,1.100),(1.071,1.099),(1.069,1.101),(1.019,1.041)],
        "NL": [(1.296,1.410),(1.304,1.450),(1.293,1.475),(1.175,1.241)],
        "NO": [(1.010,1.010),(1.009,1.013),(1.010,1.010),(1.004,1.010)],
        "PL": [(1.033,1.047),(1.032,1.056),(1.059,1.071),(1.010,1.028)],
        "PT": [(1.072,1.130),(1.111,1.185),(1.055,1.133),(1.155,1.209)],
        "RO": [(1.254,1.334),(1.306,1.366),(1.232,1.364),(1.218,1.312)],
        "SE": [(1.084,1.100),(1.072,1.086),(1.023,1.055),(1.017,1.021)],
        "SI": [(1.011,1.019),(1.014,1.026),(1.020,1.026),(1.015,1.023)],
        "SK": [(1.007,1.011),(1.005,1.011),(1.007,1.011),(1.007,1.015)],
    },
    "0.10": {
        "AT": [(1.122,1.214),(1.123,1.163),(1.117,1.133),(1.111,1.147)],
        "BE": [(1.254,1.312),(1.293,1.373),(1.327,1.399),(1.243,1.353)],
        "BG": [(1.349,1.421),(1.266,1.326),(1.259,1.333),(1.206,1.278)],
        "CH": [(1.056,1.070),(1.054,1.062),(1.056,1.068),(1.077,1.093)],
        "CZ": [(1.127,1.155),(1.110,1.140),(1.169,1.229),(1.266,1.318)],
        "DE": [(1.250,1.306),(1.279,1.349),(1.271,1.375),(1.253,1.361)],
        "DK": [(1.100,1.146),(1.125,1.195),(1.073,1.137),(1.054,1.080)],
        "EE": [(1.025,1.047),(1.020,1.032),(1.014,1.024),None],
        "ES": [(1.085,1.113),(1.129,1.203),(1.095,1.189),(1.085,1.161)],
        "FI": [(1.009,1.015),(1.007,1.011),(1.005,1.011),(0.999,1.003)],
        "FR": [(1.140,1.206),(1.148,1.190),(1.210,1.354),(1.124,1.176)],
        "GB": [(1.175,1.223),(1.345,1.397),(1.200,1.328),(1.105,1.155)],
        "GR": [(1.174,1.226),(1.175,1.201),(1.193,1.281),(1.172,1.222)],
        "HR": [(1.109,1.147),(1.122,1.216),(1.126,1.208),(1.181,1.225)],
        "HU": [(1.098,1.182),(1.054,1.084),(1.074,1.112),(1.059,1.141)],
        "IE": [None, None, None, (0.999,1.001)],
        "IT": [(1.100,1.124),(1.165,1.198),(1.200,1.276),(1.137,1.163)],
        "LT": [(1.204,1.272),(1.150,1.248),(1.129,1.248),(1.105,1.153)],
        "LU": [(1.239,1.297),(1.221,1.285),(1.312,1.365),(1.360,1.448)],
        "LV": [(1.044,1.066),(1.044,1.084),(1.025,1.047),(1.014,1.020)],
        "NL": [(1.191,1.261),(1.192,1.248),(1.173,1.219),(1.122,1.178)],
        "NO": [(1.007,1.011),None,(1.000,1.008),None],
        "PL": [(1.009,1.029),(1.011,1.021),(1.039,1.055),None],
        "PT": [(1.043,1.065),(1.057,1.133),(1.049,1.071),(1.051,1.133)],
        "RO": [(1.162,1.228),(1.202,1.306),(1.171,1.223),(1.166,1.232)],
        "SE": [(1.089,1.115),(1.074,1.096),(1.018,1.034),(1.014,1.020)],
        "SI": [(1.009,1.013),(1.010,1.016),(1.018,1.030),(1.012,1.020)],
        "SK": [(1.004,1.010),(1.007,1.011),None,(1.009,1.013)],
    },
}

# ---------------------------------------------------------------------------
# Build flat DataFrame
# ---------------------------------------------------------------------------
rows = []
for eps, cdata in RAW.items():
    for country in COUNTRIES:
        for i, date in enumerate(DATES):
            ci = cdata[country][i]
            if ci is None:
                lo, hi, mid, wid = None, None, None, None
            else:
                lo, hi = ci
                mid = round((lo + hi) / 2, 4)
                wid = round(hi - lo, 4)
            rows.append({
                "Epsilon": eps,
                "Country": country,
                "Date": date,
                "Lower": lo,
                "Upper": hi,
                "Midpoint": mid,
                "Width": wid,
                "Focus": country in FOCUS,
            })

df = pd.DataFrame(rows)

# ---------------------------------------------------------------------------
# Excel build
# ---------------------------------------------------------------------------
wb = Workbook()

# ---- Colour palette --------------------------------------------------------
HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SUB_FILL   = PatternFill("solid", fgColor="2F5496")   # mid blue
FOCUS_FILL = PatternFill("solid", fgColor="FFF2CC")   # pale yellow
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")   # very light blue
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
SUB_FONT   = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(size=10)
FOCUS_FONT = Font(bold=True, size=10)
DEGENERATE = "—"

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def fmt_num(ws, cell, val):
    if val is None:
        cell.value = DEGENERATE
        cell.alignment = Alignment(horizontal="center")
    else:
        cell.value = val
        cell.number_format = "0.000"
        cell.alignment = Alignment(horizontal="center")

# ============================================================
# SHEET 1 — FULL PIVOT  (one sheet per epsilon)
# ============================================================
EPS_LABELS = {"0.01": "ε = 0.01", "0.05": "ε = 0.05", "0.10": "ε = 0.10"}

for eps in ["0.01", "0.05", "0.10"]:
    ws = wb.create_sheet(title=f"ε={eps}")
    sub = df[df["Epsilon"] == eps].copy()

    # ----- header row 1: date groups -----
    ws.cell(1, 1, "Country").fill  = HDR_FILL
    ws.cell(1, 1).font             = HDR_FONT
    ws.cell(1, 1).alignment        = Alignment(horizontal="center", vertical="center")

    col = 2
    for date in DATES:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)
        c = ws.cell(1, col, date)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        col += 4

    # summary columns
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)
    c = ws.cell(1, col, "Summary (across 4 dates)")
    c.fill      = SUB_FILL
    c.font      = SUB_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ----- header row 2: metrics -----
    metrics = ["Lower", "Upper", "Midpoint", "Width"]
    ws.cell(2, 1, f"{EPS_LABELS[eps]}").fill  = HDR_FILL
    ws.cell(2, 1).font                         = HDR_FONT
    ws.cell(2, 1).alignment                    = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    col = 2
    for _ in DATES:
        for m in metrics:
            c = ws.cell(2, col, m)
            c.fill      = HDR_FILL
            c.font      = HDR_FONT
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            col += 1

    summary_metrics = ["Min midpoint", "Max midpoint", "Mean midpoint", "Max width"]
    for sm in summary_metrics:
        c = ws.cell(2, col, sm)
        c.fill      = SUB_FILL
        c.font      = SUB_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        col += 1

    # ----- data rows -----
    for r_idx, country in enumerate(COUNTRIES):
        row = r_idx + 3
        crow = sub[sub["Country"] == country]

        is_focus = country in FOCUS
        fill     = FOCUS_FILL if is_focus else (ALT_FILL if r_idx % 2 == 0 else WHITE_FILL)
        font     = FOCUS_FONT if is_focus else BODY_FONT

        c = ws.cell(row, 1, country)
        c.fill      = fill
        c.font      = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

        midpoints = []
        widths    = []
        col = 2
        for date in DATES:
            cell_data = crow[crow["Date"] == date].iloc[0]
            for m in metrics:
                cell = ws.cell(row, col)
                val  = cell_data[m]
                fmt_num(ws, cell, val)
                cell.fill   = fill
                cell.font   = font
                cell.border = thin_border()
                if val is not None:
                    if m == "Midpoint":
                        midpoints.append(val)
                    if m == "Width":
                        widths.append(val)
                col += 1

        # summary
        def sfmt(ws, row, col, val, fill, font):
            cell = ws.cell(row, col)
            if val is None:
                cell.value = DEGENERATE
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value         = round(val, 4)
                cell.number_format = "0.000"
                cell.alignment     = Alignment(horizontal="center")
            cell.fill   = fill
            cell.font   = font
            cell.border = thin_border()

        sfmt(ws, row, col,   min(midpoints) if midpoints else None, fill, font)
        sfmt(ws, row, col+1, max(midpoints) if midpoints else None, fill, font)
        sfmt(ws, row, col+2, round(sum(midpoints)/len(midpoints), 4) if midpoints else None, fill, font)
        sfmt(ws, row, col+3, max(widths) if widths else None, fill, font)

    # ----- column widths -----
    ws.column_dimensions["A"].width = 11
    for c in range(2, col + 4):
        ws.column_dimensions[get_column_letter(c)].width = 9
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "B3"

# ============================================================
# SHEET 2 — FOCUS COUNTRIES  (GR, ES, CH, PT) all epsilons
# ============================================================
ws2 = wb.create_sheet(title="Focus Countries")
ws2.cell(1, 1, "Focus: GR, ES, CH, PT — CI midpoints and widths across all ε").font = Font(bold=True, size=12)
ws2.merge_cells("A1:R1")

# header
headers = ["Country", "ε"] + [f"{d}\nMidpoint" for d in DATES] + \
          [f"{d}\nWidth" for d in DATES] + ["Mean\nMidpoint", "Max\nWidth"]
for ci, h in enumerate(headers, 1):
    cell = ws2.cell(2, ci, h)
    cell.fill      = HDR_FILL
    cell.font      = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()
ws2.row_dimensions[2].height = 32

r = 3
for country in ["GR", "ES", "CH", "PT"]:
    for eps in ["0.01", "0.05", "0.10"]:
        crow = df[(df["Country"] == country) & (df["Epsilon"] == eps)]
        fill = FOCUS_FILL if eps == "0.05" else (ALT_FILL if eps == "0.01" else WHITE_FILL)

        ws2.cell(r, 1, country).fill  = fill
        ws2.cell(r, 1).font           = FOCUS_FONT
        ws2.cell(r, 1).alignment      = Alignment(horizontal="center")
        ws2.cell(r, 1).border         = thin_border()

        ws2.cell(r, 2, f"ε={eps}").fill  = fill
        ws2.cell(r, 2).font              = BODY_FONT
        ws2.cell(r, 2).alignment         = Alignment(horizontal="center")
        ws2.cell(r, 2).border            = thin_border()

        midpoints, widths = [], []
        for di, date in enumerate(DATES):
            row_data = crow[crow["Date"] == date].iloc[0]
            mid  = row_data["Midpoint"]
            wid  = row_data["Width"]
            col_mid = 3 + di
            col_wid = 7 + di
            for col_i, val in [(col_mid, mid), (col_wid, wid)]:
                c = ws2.cell(r, col_i)
                fmt_num(ws2, c, val)
                c.fill   = fill
                c.font   = BODY_FONT
                c.border = thin_border()
            if mid is not None:
                midpoints.append(mid)
            if wid is not None:
                widths.append(wid)

        # mean midpoint, max width
        c_mean = ws2.cell(r, 11)
        c_mean.value         = round(sum(midpoints)/len(midpoints), 4) if midpoints else None
        c_mean.number_format = "0.000"
        c_mean.alignment     = Alignment(horizontal="center")
        c_mean.fill          = fill
        c_mean.font          = Font(bold=True, size=10)
        c_mean.border        = thin_border()

        c_maxw = ws2.cell(r, 12)
        c_maxw.value         = round(max(widths), 4) if widths else None
        c_maxw.number_format = "0.000"
        c_maxw.alignment     = Alignment(horizontal="center")
        c_maxw.fill          = fill
        c_maxw.font          = BODY_FONT
        c_maxw.border        = thin_border()

        r += 1
    # blank separator row between countries
    r += 1

# column widths
ws2.column_dimensions["A"].width = 11
ws2.column_dimensions["B"].width = 9
for c in range(3, 13):
    ws2.column_dimensions[get_column_letter(c)].width = 11
ws2.row_dimensions[1].height = 22
ws2.freeze_panes = "C3"

# ============================================================
# SHEET 3 — FLAT TABLE (all countries / all dates / all ε)
# ============================================================
ws3 = wb.create_sheet(title="Flat Data")
flat_headers = ["Epsilon", "Country", "Focus?", "Date", "Lower", "Upper", "Midpoint", "Width"]
for ci, h in enumerate(flat_headers, 1):
    cell = ws3.cell(1, ci, h)
    cell.fill      = HDR_FILL
    cell.font      = HDR_FONT
    cell.alignment = Alignment(horizontal="center")
    cell.border    = thin_border()

for r_idx, row_data in enumerate(df.itertuples(), 2):
    is_focus = row_data.Focus
    fill = FOCUS_FILL if is_focus else (ALT_FILL if r_idx % 2 == 0 else WHITE_FILL)
    vals = [row_data.Epsilon, row_data.Country,
            "Yes" if is_focus else "", row_data.Date,
            row_data.Lower, row_data.Upper, row_data.Midpoint, row_data.Width]
    for ci, val in enumerate(vals, 1):
        c = ws3.cell(r_idx, ci)
        if isinstance(val, float) and val is not None:
            c.value         = val
            c.number_format = "0.000"
            c.alignment     = Alignment(horizontal="center")
        elif val is None:
            c.value     = DEGENERATE
            c.alignment = Alignment(horizontal="center")
        else:
            c.value     = val
            c.alignment = Alignment(horizontal="center")
        c.fill   = fill
        c.font   = FOCUS_FONT if is_focus else BODY_FONT
        c.border = thin_border()

col_widths = [8, 10, 8, 9, 9, 9, 10, 9]
for ci, w in enumerate(col_widths, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.freeze_panes = "A2"

# ---- remove default empty sheet ----
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

out_path = r"C:\Users\montr\OneDrive - University of Southampton\wind_pv_synergy\results\ci_midpoints_storage.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")

# ---- print focus country summary to console ----
print("\nFocus country midpoints at ε=0.05:")
focus_df = df[(df["Focus"]) & (df["Epsilon"] == "0.05")][["Country","Date","Midpoint","Width"]]
print(focus_df.to_string(index=False))
print(f"\nAll ε=0.05 focus midpoints — min: {focus_df['Midpoint'].min():.3f}, max: {focus_df['Midpoint'].max():.3f}")
