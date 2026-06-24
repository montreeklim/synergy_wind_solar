#!/usr/bin/env python3
"""
Comparison table: mean synergy ratio on day 271 across battery storage sizes
(50%, 100%, 200% of installed capacity) and tolerance values (0.01, 0.05, 0.10).

Reads raw per-country CSVs from:
  - 50%  : ../battery_results/50_installed_capacity/battery_{COUNTRY}_day_271_tol_{TOL}_50_installed_capacity.csv
  - 100% : ../battery_results/100_installed_capacity/battery_{COUNTRY}_day_271_tol_{TOL}_100_installed_capacity.csv
  - 200% : ../battery_results/200_installed_capacity/battery_{COUNTRY}_day_271_tol_{TOL}_200_installed_capacity.csv

Outputs:
  ../results/storage_size_comparison_day271.csv   — wide table (countries × 9 columns)
  ../results/storage_size_comparison_day271.xlsx  — same with formatting
"""

import os
import pandas as pd
import numpy as np

# ── paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT       = os.path.dirname(SCRIPT_DIR)
RESULTS    = os.path.join(ROOT, "results")
os.makedirs(RESULTS, exist_ok=True)

# ── country list (canonical order) ────────────────────────────────────────────
COUNTRIES = [
    "AT", "BE", "BG", "CH", "CZ", "DE", "DK", "EE", "ES", "FI",
    "FR", "EL", "HR", "HU", "IE", "IT", "LT", "LU", "LV", "NL",
    "NO", "PL", "PT", "RO", "SI", "SK", "SE", "UK",
]

TOLERANCES = [0.01, 0.05, 0.10]
SIZES      = [50, 100, 200]   # % of installed capacity


def tol_str(tol: float) -> str:
    """Return the tolerance string as it appears in result filenames (e.g. 0.1, not 0.10)."""
    return str(tol)


def load_mean_ratio(country: str, tol: float, size: int) -> float:
    """Return mean ratio across all sets for one (country, tol, size) triple."""
    tol_s = tol_str(tol)
    path = os.path.join(
        ROOT, "battery_results", f"{size}_installed_capacity",
        f"battery_{country}_day_271_tol_{tol_s}_{size}_installed_capacity.csv",
    )

    if not os.path.exists(path):
        return np.nan

    df = pd.read_csv(path)
    # column may be 'ratio' — keep only finite, positive values
    ratios = pd.to_numeric(df["ratio"], errors="coerce")
    ratios = ratios[np.isfinite(ratios) & (ratios > 0)]
    return round(ratios.mean(), 3) if len(ratios) > 0 else np.nan


# ── build table ───────────────────────────────────────────────────────────────
records = []
for country in COUNTRIES:
    row = {"Country": country}
    for tol in TOLERANCES:
        for size in SIZES:
            col = f"ε={tol_str(tol)} / {size}%"
            row[col] = load_mean_ratio(country, tol, size)
    records.append(row)

df_out = pd.DataFrame(records).set_index("Country")

# ── CSV output ────────────────────────────────────────────────────────────────
csv_path = os.path.join(RESULTS, "storage_size_comparison_day271.csv")
df_out.to_csv(csv_path)
print(f"Saved: {csv_path}")

# ── Excel output ──────────────────────────────────────────────────────────────
xlsx_path = os.path.join(RESULTS, "storage_size_comparison_day271.xlsx")

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    df_out.to_excel(writer, sheet_name="Day271_Storage_Comparison")
    ws = writer.sheets["Day271_Storage_Comparison"]

    # ── formatting helpers ────────────────────────────────────────────────────
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill  = PatternFill("solid", fgColor="1F4E79")   # dark blue
    header_font  = Font(color="FFFFFF", bold=True)
    group_fills  = [
        PatternFill("solid", fgColor="D9E1F2"),  # light blue  — tol 0.01
        PatternFill("solid", fgColor="E2EFDA"),  # light green — tol 0.05
        PatternFill("solid", fgColor="FFF2CC"),  # light amber — tol 0.10
    ]
    thin_side    = Side(style="thin")
    thick_side   = Side(style="medium")

    def thin_border(left=False, right=False, top=False, bottom=False):
        return Border(
            left   = thick_side if left   else thin_side,
            right  = thick_side if right  else thin_side,
            top    = thick_side if top    else thin_side,
            bottom = thick_side if bottom else thin_side,
        )

    n_rows = len(COUNTRIES) + 2   # header row 1 (tol group) + header row 2 (country/size) + data
    n_cols = len(TOLERANCES) * len(SIZES) + 1   # country col + 9 data cols

    # Insert a row above the existing header for tolerance group labels
    ws.insert_rows(1)

    # Row 1: tolerance group spans  (cols B–D = tol0.01, E–G = tol0.05, H–J = tol0.10)
    group_start_cols = [2, 5, 8]   # 1-indexed; col 1 is "Country"
    group_labels = ["ε = 0.01", "ε = 0.05", "ε = 0.10"]
    for gi, (gcol, glabel) in enumerate(zip(group_start_cols, group_labels)):
        cell = ws.cell(row=1, column=gcol, value=glabel)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(
            start_row=1, start_column=gcol,
            end_row=1,   end_column=gcol + len(SIZES) - 1,
        )

    # Row 2: country label + size sub-headers
    ws.cell(row=2, column=1, value="Country").fill = header_fill
    ws.cell(row=2, column=1).font = header_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    col_idx = 2
    for gi, tol in enumerate(TOLERANCES):
        for size in SIZES:
            cell = ws.cell(row=2, column=col_idx, value=f"{size}%")
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            col_idx += 1

    # Data rows (start at row 3)
    for r_offset, country in enumerate(COUNTRIES):
        r = r_offset + 3
        ws.cell(row=r, column=1, value=country).font = Font(bold=True)
        col_idx = 2
        for gi, tol in enumerate(TOLERANCES):
            for size in SIZES:
                col_name = f"ε={tol_str(tol)} / {size}%"
                val = df_out.loc[country, col_name]
                cell = ws.cell(row=r, column=col_idx)
                if pd.isna(val):
                    cell.value = "—"
                else:
                    cell.value = val
                    cell.number_format = "0.000"
                cell.fill = group_fills[gi]
                cell.alignment = Alignment(horizontal="center")
                col_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 10
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16

print(f"Saved: {xlsx_path}")

# ── console preview ───────────────────────────────────────────────────────────
print("\nMean synergy ratio — Day 271, by storage size and tolerance")
print("=" * 80)
print(df_out.to_string(na_rep="—"))
