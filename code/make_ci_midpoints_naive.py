"""
Generate CI midpoints/widths Excel for naive (no-storage) rolling-window results.
Data sourced from results/naive_CI_day{DAY}_tol{TAG}.csv (matches
writing/section_6_absolute_profits.tex tables tab:naive_e01/e05/e10).
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
RESULTS_DIR = r"C:\Users\montr\OneDrive - University of Southampton\wind_pv_synergy\results"

DATES = ["D-271", "D-301", "D-332", "D-362"]
DAY_BY_DATE = {"D-271": 271, "D-301": 301, "D-332": 332, "D-362": 362}
TOL_TAG = {"0.01": "01", "0.05": "05", "0.10": "10"}

COUNTRIES = [
    "AT","BE","BG","CH","CZ","DE","DK","EE","ES","FI",
    "FR","GB","GR","HR","HU","IE","IT","LT","LU","LV",
    "NL","NO","PL","PT","RO","SE","SI","SK",
]

FOCUS = {"GR", "ES", "CH", "PT"}

# Display codes (GB/GR, matching the paper and ci_midpoints_storage.xlsx)
# vs. the ISO codes used in naive_CI_*.csv (UK/EL).
CSV_COUNTRY = {"GB": "UK", "GR": "EL"}

# ---------------------------------------------------------------------------
# Load CI data from naive_CI_day{DAY}_tol{TAG}.csv
# ---------------------------------------------------------------------------
rows = []
for eps, tag in TOL_TAG.items():
    for date in DATES:
        day = DAY_BY_DATE[date]
        csv_path = fr"{RESULTS_DIR}\naive_CI_day{day}_tol{tag}.csv"
        ci = pd.read_csv(csv_path, index_col="country")
        for country in COUNTRIES:
            csv_country = CSV_COUNTRY.get(country, country)
            lo_raw = ci.loc[csv_country, "ci_lower"]
            hi_raw = ci.loc[csv_country, "ci_upper"]
            if lo_raw == "---":
                lo, hi, mid, wid = None, None, None, None
            else:
                lo, hi = float(lo_raw), float(hi_raw)
                mid = round((lo + hi) / 2, 4)
                wid = round(hi - lo, 4)
            rows.append({
                "Epsilon": eps,
                "Country": country,
                "Date": date,
                "Lower": lo,
                "Upper": hi,
                "Midpoint": mid,
                "Width": wid,
                "Focus": country in FOCUS,
            })

df = pd.DataFrame(rows)

# ---------------------------------------------------------------------------
# Excel build
# ---------------------------------------------------------------------------
wb = Workbook()

# ---- Colour palette --------------------------------------------------------
HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SUB_FILL   = PatternFill("solid", fgColor="2F5496")   # mid blue
FOCUS_FILL = PatternFill("solid", fgColor="FFF2CC")   # pale yellow
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")   # very light blue
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
SUB_FONT   = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(size=10)
FOCUS_FONT = Font(bold=True, size=10)
DEGENERATE = "—"

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def fmt_num(ws, cell, val):
    if val is None or pd.isna(val):
        cell.value = DEGENERATE
        cell.alignment = Alignment(horizontal="center")
    else:
        cell.value = val
        cell.number_format = "0.000"
        cell.alignment = Alignment(horizontal="center")

# ============================================================
# SHEET 1 — FULL PIVOT  (one sheet per epsilon)
# ============================================================
EPS_LABELS = {"0.01": "ε = 0.01", "0.05": "ε = 0.05", "0.10": "ε = 0.10"}

for eps in ["0.01", "0.05", "0.10"]:
    ws = wb.create_sheet(title=f"ε={eps}")
    sub = df[df["Epsilon"] == eps].copy()

    # ----- header row 1: date groups -----
    ws.cell(1, 1, "Country").fill  = HDR_FILL
    ws.cell(1, 1).font             = HDR_FONT
    ws.cell(1, 1).alignment        = Alignment(horizontal="center", vertical="center")

    col = 2
    for date in DATES:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)
        c = ws.cell(1, col, date)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        col += 4

    # summary columns
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)
    c = ws.cell(1, col, "Summary (across 4 dates)")
    c.fill      = SUB_FILL
    c.font      = SUB_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ----- header row 2: metrics -----
    metrics = ["Lower", "Upper", "Midpoint", "Width"]
    ws.cell(2, 1, f"{EPS_LABELS[eps]}").fill  = HDR_FILL
    ws.cell(2, 1).font                         = HDR_FONT
    ws.cell(2, 1).alignment                    = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    col = 2
    for _ in DATES:
        for m in metrics:
            c = ws.cell(2, col, m)
            c.fill      = HDR_FILL
            c.font      = HDR_FONT
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            col += 1

    summary_metrics = ["Min midpoint", "Max midpoint", "Mean midpoint", "Max width"]
    for sm in summary_metrics:
        c = ws.cell(2, col, sm)
        c.fill      = SUB_FILL
        c.font      = SUB_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        col += 1

    # ----- data rows -----
    for r_idx, country in enumerate(COUNTRIES):
        row = r_idx + 3
        crow = sub[sub["Country"] == country]

        is_focus = country in FOCUS
        fill     = FOCUS_FILL if is_focus else (ALT_FILL if r_idx % 2 == 0 else WHITE_FILL)
        font     = FOCUS_FONT if is_focus else BODY_FONT

        c = ws.cell(row, 1, country)
        c.fill      = fill
        c.font      = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

        midpoints = []
        widths    = []
        col = 2
        for date in DATES:
            cell_data = crow[crow["Date"] == date].iloc[0]
            for m in metrics:
                cell = ws.cell(row, col)
                val  = cell_data[m]
                fmt_num(ws, cell, val)
                cell.fill   = fill
                cell.font   = font
                cell.border = thin_border()
                if val is not None and not pd.isna(val):
                    if m == "Midpoint":
                        midpoints.append(val)
                    if m == "Width":
                        widths.append(val)
                col += 1

        # summary
        def sfmt(ws, row, col, val, fill, font):
            cell = ws.cell(row, col)
            if val is None:
                cell.value = DEGENERATE
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value         = round(val, 4)
                cell.number_format = "0.000"
                cell.alignment     = Alignment(horizontal="center")
            cell.fill   = fill
            cell.font   = font
            cell.border = thin_border()

        sfmt(ws, row, col,   min(midpoints) if midpoints else None, fill, font)
        sfmt(ws, row, col+1, max(midpoints) if midpoints else None, fill, font)
        sfmt(ws, row, col+2, round(sum(midpoints)/len(midpoints), 4) if midpoints else None, fill, font)
        sfmt(ws, row, col+3, max(widths) if widths else None, fill, font)

    # ----- column widths -----
    ws.column_dimensions["A"].width = 11
    for c in range(2, col + 4):
        ws.column_dimensions[get_column_letter(c)].width = 9
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "B3"

# ============================================================
# SHEET 2 — FOCUS COUNTRIES  (GR, ES, CH, PT) all epsilons
# ============================================================
ws2 = wb.create_sheet(title="Focus Countries")
ws2.cell(1, 1, "Focus: GR, ES, CH, PT — naive model CI midpoints and widths across all ε").font = Font(bold=True, size=12)
ws2.merge_cells("A1:R1")

# header
headers = ["Country", "ε"] + [f"{d}\nMidpoint" for d in DATES] + \
          [f"{d}\nWidth" for d in DATES] + ["Mean\nMidpoint", "Max\nWidth"]
for ci, h in enumerate(headers, 1):
    cell = ws2.cell(2, ci, h)
    cell.fill      = HDR_FILL
    cell.font      = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()
ws2.row_dimensions[2].height = 32

r = 3
for country in ["GR", "ES", "CH", "PT"]:
    for eps in ["0.01", "0.05", "0.10"]:
        crow = df[(df["Country"] == country) & (df["Epsilon"] == eps)]
        fill = FOCUS_FILL if eps == "0.05" else (ALT_FILL if eps == "0.01" else WHITE_FILL)

        ws2.cell(r, 1, country).fill  = fill
        ws2.cell(r, 1).font           = FOCUS_FONT
        ws2.cell(r, 1).alignment      = Alignment(horizontal="center")
        ws2.cell(r, 1).border         = thin_border()

        ws2.cell(r, 2, f"ε={eps}").fill  = fill
        ws2.cell(r, 2).font              = BODY_FONT
        ws2.cell(r, 2).alignment         = Alignment(horizontal="center")
        ws2.cell(r, 2).border            = thin_border()

        midpoints, widths = [], []
        for di, date in enumerate(DATES):
            row_data = crow[crow["Date"] == date].iloc[0]
            mid  = row_data["Midpoint"]
            wid  = row_data["Width"]
            col_mid = 3 + di
            col_wid = 7 + di
            for col_i, val in [(col_mid, mid), (col_wid, wid)]:
                c = ws2.cell(r, col_i)
                fmt_num(ws2, c, val)
                c.fill   = fill
                c.font   = BODY_FONT
                c.border = thin_border()
            if mid is not None and not pd.isna(mid):
                midpoints.append(mid)
            if wid is not None and not pd.isna(wid):
                widths.append(wid)

        # mean midpoint, max width
        c_mean = ws2.cell(r, 11)
        c_mean.value         = round(sum(midpoints)/len(midpoints), 4) if midpoints else DEGENERATE
        if midpoints:
            c_mean.number_format = "0.000"
        c_mean.alignment     = Alignment(horizontal="center")
        c_mean.fill          = fill
        c_mean.font          = Font(bold=True, size=10)
        c_mean.border        = thin_border()

        c_maxw = ws2.cell(r, 12)
        c_maxw.value         = round(max(widths), 4) if widths else DEGENERATE
        if widths:
            c_maxw.number_format = "0.000"
        c_maxw.alignment     = Alignment(horizontal="center")
        c_maxw.fill          = fill
        c_maxw.font          = BODY_FONT
        c_maxw.border        = thin_border()

        r += 1
    # blank separator row between countries
    r += 1

# column widths
ws2.column_dimensions["A"].width = 11
ws2.column_dimensions["B"].width = 9
for c in range(3, 13):
    ws2.column_dimensions[get_column_letter(c)].width = 11
ws2.row_dimensions[1].height = 22
ws2.freeze_panes = "C3"

# ============================================================
# SHEET 3 — FLAT TABLE (all countries / all dates / all ε)
# ============================================================
ws3 = wb.create_sheet(title="Flat Data")
flat_headers = ["Epsilon", "Country", "Focus?", "Date", "Lower", "Upper", "Midpoint", "Width"]
for ci, h in enumerate(flat_headers, 1):
    cell = ws3.cell(1, ci, h)
    cell.fill      = HDR_FILL
    cell.font      = HDR_FONT
    cell.alignment = Alignment(horizontal="center")
    cell.border    = thin_border()

for r_idx, row_data in enumerate(df.itertuples(), 2):
    is_focus = row_data.Focus
    fill = FOCUS_FILL if is_focus else (ALT_FILL if r_idx % 2 == 0 else WHITE_FILL)
    vals = [row_data.Epsilon, row_data.Country,
            "Yes" if is_focus else "", row_data.Date,
            row_data.Lower, row_data.Upper, row_data.Midpoint, row_data.Width]
    for ci, val in enumerate(vals, 1):
        c = ws3.cell(r_idx, ci)
        if isinstance(val, float) and not pd.isna(val):
            c.value         = val
            c.number_format = "0.000"
            c.alignment     = Alignment(horizontal="center")
        elif isinstance(val, float) and pd.isna(val):
            c.value     = DEGENERATE
            c.alignment = Alignment(horizontal="center")
        else:
            c.value     = val
            c.alignment = Alignment(horizontal="center")
        c.fill   = fill
        c.font   = FOCUS_FONT if is_focus else BODY_FONT
        c.border = thin_border()

col_widths = [8, 10, 8, 9, 9, 9, 10, 9]
for ci, w in enumerate(col_widths, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.freeze_panes = "A2"

# ---- remove default empty sheet ----
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

out_path = r"C:\Users\montr\OneDrive - University of Southampton\wind_pv_synergy\results\ci_midpoints_naive.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")

# ---- print focus country summary to console ----
print("\nFocus country midpoints at eps=0.05:")
focus_df = df[(df["Focus"]) & (df["Epsilon"] == "0.05")][["Country","Date","Midpoint","Width"]]
print(focus_df.to_string(index=False))
print(f"\nAll eps=0.05 focus midpoints - min: {focus_df['Midpoint'].min():.3f}, max: {focus_df['Midpoint'].max():.3f}")
